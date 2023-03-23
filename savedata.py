import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from explaindata2 import explain_data

#delete_html_files_input = input("是否删除所有 HTML 文件？输入 1 表示是，输入 0 表示否：")


def save_data(hotel_name):
    file_name = f"{hotel_name}.xlsx"

    # 指定保存 Excel 文件的路径
    desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    xls_path = os.path.join(desktop_path, file_name)

    # 创建或打开Excel文件
    try:
        wb = load_workbook(filename=xls_path)
    except FileNotFoundError:
        wb = Workbook()

    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    ws_title = f"Sheet1_{current_time}"

    # 判断是否已存在该表名，如果已存在，则添加数字后缀直至不存在为止
    suffix = 1
    while ws_title in wb.sheetnames:
        ws_title = f"Sheet{suffix}_{current_time}"
        suffix += 1

    # 创建新的工作表并写入数据
    ws = wb.create_sheet(title=ws_title)
    ws.append(["Check-in", "Room Name", "Price", "Status"])

    # 为第一行添加筛选功能
    ws.auto_filter.ref = "A1:D1"

    # 解析数据，并按 Check-in 列排序
    new_data = []
    for filename in os.listdir("."):
        if filename.endswith(".html"):
            with open(filename, "r", encoding="utf-8") as f:
                html_content = f.read()
                parsed_data = explain_data(html_content)

                # 将日期字符串转换为 datetime 对象
                for row in parsed_data:
                    date_string = row[0].replace("年", "-").replace("月", "-").replace("日", "")
                    row[0] = datetime.strptime(date_string, '%Y-%m-%d')
                    row[1] = [str(name) for name in row[1]]  # 更改 'Room Name' 列的格式
                new_data.extend(parsed_data)

    new_data = sorted(new_data, key=lambda row: row[0])

    for new_row_num, new_row in enumerate(new_data, start=2):
        # 将日期格式化为字符串
        new_row[0] = new_row[0].strftime('%Y-%m-%d')

        # 将包含房间名称的列表转换为字符串
        new_room_names = ", ".join(map(str, new_row[1]))

        # 写入新行数据
        ws.append([new_row[0], new_room_names, new_row[2], new_row[3]])

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # 打印 Excel 文件路径
    print(f"Excel file path: {xls_path}")

    # 保存 Excel 文件
    wb.save(filename=xls_path)

    # 根据需要删除 HTML 文件
    # while True:
    #     if delete_html_files_input == "1":
    #         delete_html_files()
    #         break
    #     elif delete_html_files_input == "0":
    #         break
    #     else:
    #         print("输入无效，请重新输入。")

    # 保存成功的消息
    print(f"Excel file saved successfully to {xls_path}.")


if __name__ == '__main__':
    save_data()
